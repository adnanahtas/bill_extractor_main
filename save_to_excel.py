import sys
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

def save_to_excel(json_file_path, output_file):
    try:
        # Normalize the file paths
        json_file_path = os.path.normpath(json_file_path)
        output_file = os.path.normpath(output_file)

        # Check if the JSON file exists
        if not os.path.exists(json_file_path):
            raise FileNotFoundError(f"JSON file not found: {json_file_path}")

        with open(json_file_path, 'r', encoding='utf-8') as f:
            parsed_data = json.load(f)

        wb = Workbook()
        ws = wb.active

        register = parsed_data.get("परिवार_रजिस्टर", {})
        df = pd.DataFrame(register.get("परिवार_सदस्य", []))
        # Write the header information (excluding परिवार_सदस्य)
        header_info = {k: v for k, v in register.items() if k != "परिवार_सदस्य"}
        for i, (key, value) in enumerate(header_info.items(), start=1):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=value)

        start_row = len(header_info) + 2

        # Add column headers
        for col, header in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Add data
        for r, row in enumerate(df.values, start=start_row + 1):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)

        wb.save(output_file)
        print(f"Excel file saved successfully: {output_file}")

    except Exception as e:
        print(f"Error saving Excel file: {str(e)}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python save_to_excel.py <json_file_path> <output_file>")
        sys.exit(1)

    json_file_path = sys.argv[1]
    base_name = os.path.splitext(sys.argv[1])[0]  # Get the base name without extension
    output_file = f"{base_name}.xlsx"  # Add .xlsx extension to the output file name


    try:
        save_to_excel(json_file_path, output_file)
        print(f"Excel file saved successfully: {output_file}")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}", file=sys.stderr)
        sys.exit(1)